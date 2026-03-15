from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import hashlib
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.db.models import AssetType, Category, FlowType, Holding, Household, HouseholdMember, Transaction, User
from app.schemas import ImportIssue, ImportReport
from app.services.owner_links import find_unique_household_member_by_display_name


MONTH_SHEET = re.compile(r"^(?:[1-9]|1[0-2])$")


@dataclass
class ParsedTransaction:
    source_ref: str
    dedupe_hash: str
    dedupe_ordinal: int
    occurred_on: date
    flow_type: FlowType
    major: str
    minor: str
    amount: Decimal
    memo: str
    owner_name: str | None


@dataclass
class ParsedHolding:
    asset_type: AssetType
    symbol: str
    market_symbol: str
    name: str
    category: str
    owner_name: str | None
    account_name: str | None
    quantity: Decimal
    average_cost: Decimal
    currency: str
    source_ref: str


class WorkbookImporter:
    def load_default_path(self, root_dir: Path) -> Path:
        matches = sorted((root_dir / "legacy").glob("*.xlsx"))
        if not matches:
            raise FileNotFoundError("legacy/*.xlsx not found")
        return matches[0]

    def run(
        self,
        db: Session,
        *,
        household: Household,
        workbook_path: Path,
        mode: str,
        commit: bool = True,
    ) -> ImportReport:
        wb = load_workbook(
            workbook_path,
            data_only=False,
            read_only=settings.import_read_only_mode,
        )
        try:
            if len(wb.worksheets) > settings.import_max_sheets:
                raise ValueError("workbook has too many sheets")
            metrics = self._collect_metrics(wb)
            categories, category_issues = self._parse_categories(wb)
            tx_rows, tx_issues = self._parse_transactions(wb)
            holding_rows, holding_issues = self._parse_holdings(wb)

            applied_categories = 0
            applied_transactions = 0
            skipped_transactions = 0
            applied_holdings_added = 0
            applied_holdings_updated = 0

            if mode == "apply":
                category_map, inserted = self._upsert_categories(db, household.id, categories)
                applied_categories += inserted
                tx_added, tx_skipped = self._apply_transactions(db, household.id, category_map, tx_rows)
                applied_transactions += tx_added
                skipped_transactions += tx_skipped
                h_added, h_updated, h_issues = self._apply_holdings(db, household.id, holding_rows)
                applied_holdings_added += h_added
                applied_holdings_updated += h_updated
                holding_issues.extend(h_issues)
                if commit:
                    db.commit()

            issues = metrics["issues"] + category_issues + tx_issues + holding_issues
            return ImportReport(
                workbook_path=str(workbook_path),
                sheets=int(metrics["sheets"]),
                formula_cells=int(metrics["formula_cells"]),
                merged_ranges=int(metrics["merged_ranges"]),
                chart_count=int(metrics["chart_count"]),
                monthly_formula_mismatch_count=int(metrics["mismatch_count"]),
                detected_mismatch_cells=metrics["mismatch_cells"],
                category_rows=len(categories),
                transaction_rows=len(tx_rows),
                holding_rows=len(holding_rows),
                applied_categories=applied_categories,
                applied_transactions=applied_transactions,
                applied_holdings_added=applied_holdings_added,
                applied_holdings_updated=applied_holdings_updated,
                skipped_transactions=skipped_transactions,
                issues=issues,
            )
        finally:
            wb.close()

    def _collect_metrics(self, wb) -> dict:
        sheets = len(wb.worksheets)
        formula_cells = 0
        merged_ranges = 0
        chart_count = 0
        issues: list[ImportIssue] = []

        month_formula_map: dict[str, dict[str, str]] = {}
        for ws in wb.worksheets:
            merged_ranges += len(getattr(getattr(ws, "merged_cells", None), "ranges", []) or [])
            chart_count += len(getattr(ws, "_charts", []) or [])
            by_coord: dict[str, str] = {}
            row_limit = self._sheet_row_upper_bound(ws, settings.import_max_rows_per_sheet)
            col_limit = self._sheet_col_upper_bound(ws, settings.import_max_columns_per_sheet)
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=row_limit, max_col=col_limit),
                start=1,
            ):
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        formula_cells += 1
                        if MONTH_SHEET.match(ws.title):
                            if cell.coordinate:
                                coord = cell.coordinate
                            else:
                                col = cell.column
                                col_ref = get_column_letter(col) if isinstance(col, int) else str(col)
                                coord = f"{col_ref}{row_idx}"
                            by_coord[coord] = value.strip()
            if by_coord:
                month_formula_map[ws.title] = by_coord

        mismatch_cells: list[str] = []
        if month_formula_map:
            coords = sorted(set().union(*(set(data.keys()) for data in month_formula_map.values())))
            for coord in coords:
                values = {}
                for month, mapping in month_formula_map.items():
                    if coord in mapping:
                        values.setdefault(mapping[coord], []).append(month)
                if len(values) > 1:
                    mismatch_cells.append(coord)
            if mismatch_cells:
                issues.append(
                    ImportIssue(
                        severity="warning",
                        code="MONTH_FORMULA_MISMATCH",
                        message="월별 시트에서 동일 좌표 수식이 불일치합니다.",
                        detail={
                            "mismatch_count": len(mismatch_cells),
                            "sample": mismatch_cells[:30],
                        },
                    )
                )
            if "2" in month_formula_map:
                issues.append(
                    ImportIssue(
                        severity="info",
                        code="FEB_RANGE_OUTLIER_RISK",
                        message="2월 시트의 고정 범위 수식 단축(197/198) 가능성이 있어 합계 편차를 검증해야 합니다.",
                        sheet="2",
                    )
                )

        return {
            "sheets": sheets,
            "formula_cells": formula_cells,
            "merged_ranges": merged_ranges,
            "chart_count": chart_count,
            "mismatch_count": len(mismatch_cells),
            "mismatch_cells": mismatch_cells[:200],
            "issues": issues,
        }

    def _parse_categories(self, wb) -> tuple[list[tuple[FlowType, str, str]], list[ImportIssue]]:
        rows: list[tuple[FlowType, str, str]] = []
        issues: list[ImportIssue] = []
        if "가계부 분류" not in wb.sheetnames:
            issues.append(
                ImportIssue(
                    severity="warning",
                    code="CATEGORY_SHEET_MISSING",
                    message="'가계부 분류' 시트를 찾지 못해 거래에서 파생 생성합니다.",
                    sheet="가계부 분류",
                )
            )
            return rows, issues

        ws = wb["가계부 분류"]
        seen: set[tuple[FlowType, str, str]] = set()
        upper = self._sheet_row_upper_bound_with_issue(ws, issues=issues, sheet_name=ws.title)
        for row_idx, row_values in self._iter_rows_values(ws, min_row=5, max_row=upper, max_col=4):
            major = self._text(self._row_value(row_values, 3))
            minor = self._text(self._row_value(row_values, 4))
            if not major or not minor:
                continue
            flow_type = self._guess_flow_type(major)
            item = (flow_type, major, minor)
            if item in seen:
                continue
            seen.add(item)
            rows.append(item)
        return rows, issues

    def _parse_transactions(self, wb) -> tuple[list[ParsedTransaction], list[ImportIssue]]:
        rows: list[ParsedTransaction] = []
        issues: list[ImportIssue] = []
        signature_counts: dict[str, int] = {}
        for ws in wb.worksheets:
            if not MONTH_SHEET.match(ws.title):
                continue
            upper = self._sheet_row_upper_bound_with_issue(ws, issues=issues, sheet_name=ws.title)
            for row_idx, row_values in self._iter_rows_values(ws, min_row=10, max_row=upper, max_col=6):
                occurred_on = self._to_date(self._row_value(row_values, 2))
                major = self._text(self._row_value(row_values, 3))
                minor = self._text(self._row_value(row_values, 4))
                memo = self._text(self._row_value(row_values, 5)) or ""
                amount = self._to_decimal(self._row_value(row_values, 6))
                if occurred_on is None and (major or minor or memo or amount):
                    issues.append(
                        ImportIssue(
                            severity="warning",
                            code="TX_DATE_MISSING",
                            message="날짜가 없어 거래를 건너뜁니다.",
                            sheet=ws.title,
                            row=row_idx,
                        )
                    )
                    continue
                if occurred_on is None or not major or amount is None or amount <= 0:
                    continue

                flow_type = self._guess_flow_type(major)
                owner_name = self._extract_owner_name(major, minor, memo)
                normalized_minor = minor or "기타"
                dedupe_hash = self._transaction_dedupe_hash(
                    occurred_on=occurred_on,
                    flow_type=flow_type,
                    major=major,
                    minor=normalized_minor,
                    amount=amount,
                    memo=memo,
                    owner_name=owner_name,
                )
                dedupe_ordinal = int(signature_counts.get(dedupe_hash, 0)) + 1
                signature_counts[dedupe_hash] = dedupe_ordinal
                rows.append(
                    ParsedTransaction(
                        source_ref=f"excel:tx:{dedupe_hash}:{dedupe_ordinal}",
                        dedupe_hash=dedupe_hash,
                        dedupe_ordinal=dedupe_ordinal,
                        occurred_on=occurred_on,
                        flow_type=flow_type,
                        major=major,
                        minor=normalized_minor,
                        amount=amount,
                        memo=memo,
                        owner_name=owner_name,
                    )
                )
        return rows, issues

    def _parse_holdings(self, wb) -> tuple[list[ParsedHolding], list[ImportIssue]]:
        rows: list[ParsedHolding] = []
        issues: list[ImportIssue] = []
        rows.extend(self._parse_stock_holdings(wb, issues))
        rows.extend(self._parse_cash_holdings(wb, issues))
        rows.extend(self._parse_pension_holdings(wb, issues))
        rows.extend(self._parse_real_estate_holdings(wb, issues))
        return rows, issues

    def _parse_stock_holdings(self, wb, issues: list[ImportIssue]) -> list[ParsedHolding]:
        data: list[ParsedHolding] = []
        name = "2) 주식투자"
        if name not in wb.sheetnames:
            return data
        ws = wb[name]
        upper = self._sheet_row_upper_bound_with_issue(ws, issues=issues, sheet_name=name)
        for row_idx, row_values in self._iter_rows_values(ws, min_row=7, max_row=upper, max_col=9):
            market = self._text(self._row_value(row_values, 2))
            account = self._text(self._row_value(row_values, 3))
            category = self._text(self._row_value(row_values, 4)) or "주식"
            item_name = self._text(self._row_value(row_values, 5))
            code = self._text(self._row_value(row_values, 6))
            avg_cost = self._to_decimal(self._row_value(row_values, 7))
            quantity = self._to_decimal(self._row_value(row_values, 8))
            invested = self._to_decimal(self._row_value(row_values, 9))

            if not code or quantity is None or quantity <= 0:
                continue
            symbol = self._normalize_stock_symbol(code, market)
            used_invested = False
            # Prefer explicit unit buy price from the workbook.
            if avg_cost and avg_cost > 0:
                average_cost = avg_cost.quantize(Decimal("0.0001"))
            elif invested and invested > 0:
                average_cost = (invested / quantity).quantize(Decimal("0.0001"))
                used_invested = True
            else:
                issues.append(
                    ImportIssue(
                        severity="warning",
                        code="HOLDING_COST_MISSING",
                        message="평균단가/매수금액이 없어 주식 자산을 건너뜁니다.",
                        sheet=name,
                        row=row_idx,
                    )
                )
                continue
            currency = "KRW" if symbol.endswith(".KR") else "USD"
            # "매수 금액(\\)" column is KRW total; if used as fallback, keep KRW basis.
            if used_invested:
                currency = "KRW"
            owner = self._extract_owner_name(account, item_name, item_name)
            data.append(
                ParsedHolding(
                    asset_type=AssetType.stock,
                    symbol=symbol,
                    market_symbol=symbol,
                    name=item_name or symbol,
                    category=category,
                    owner_name=owner,
                    account_name=account,
                    quantity=quantity,
                    average_cost=average_cost,
                    currency=currency,
                    source_ref=f"excel:{name}:{row_idx}",
                )
            )
        return data

    def _parse_cash_holdings(self, wb, issues: list[ImportIssue]) -> list[ParsedHolding]:
        data: list[ParsedHolding] = []
        name = "3) 저축 및 현금성"
        if name not in wb.sheetnames:
            return data
        ws = wb[name]
        group = "현금성"
        upper = self._sheet_row_upper_bound_with_issue(ws, issues=issues, sheet_name=name)
        for row_idx, row_values in self._iter_rows_values(ws, min_row=7, max_row=upper, max_col=8):
            val_group = self._text(self._row_value(row_values, 2))
            if val_group:
                group = val_group
            item_name = self._text(self._row_value(row_values, 3))
            bank = self._text(self._row_value(row_values, 4))
            account = self._text(self._row_value(row_values, 5))
            amount = self._to_decimal(self._row_value(row_values, 8))
            if not item_name or amount is None or amount <= 0:
                continue
            symbol = self._build_cash_symbol(item_name, account, row_idx, "CASH")
            owner = self._extract_owner_name(group, item_name, item_name)
            account_name = " / ".join([part for part in [bank, account] if part]) or bank
            data.append(
                ParsedHolding(
                    asset_type=AssetType.cash,
                    symbol=symbol,
                    market_symbol=symbol,
                    name=" / ".join([part for part in [group, item_name, bank] if part]),
                    category=group,
                    owner_name=owner,
                    account_name=account_name,
                    quantity=Decimal("1"),
                    average_cost=amount.quantize(Decimal("0.0001")),
                    currency="KRW",
                    source_ref=f"excel:{name}:{row_idx}",
                )
            )
        return data

    def _parse_pension_holdings(self, wb, issues: list[ImportIssue]) -> list[ParsedHolding]:
        data: list[ParsedHolding] = []
        name = "4) 연금"
        if name not in wb.sheetnames:
            return data
        ws = wb[name]
        group = "연금"
        upper = self._sheet_row_upper_bound_with_issue(ws, issues=issues, sheet_name=name)
        for row_idx, row_values in self._iter_rows_values(ws, min_row=7, max_row=upper, max_col=9):
            val_group = self._text(self._row_value(row_values, 2))
            if val_group and "합계" not in val_group:
                group = val_group
            company = self._text(self._row_value(row_values, 3))
            item_type = self._text(self._row_value(row_values, 4))
            item_name = self._text(self._row_value(row_values, 5))
            code = self._text(self._row_value(row_values, 6))
            qty = self._to_decimal(self._row_value(row_values, 8))
            principal = self._to_decimal(self._row_value(row_values, 9))
            if principal is None or principal <= 0:
                continue

            owner = self._extract_owner_name(group, item_name, item_name)
            if item_type and "ETF" in item_type and code and qty and qty > 0:
                symbol = self._normalize_stock_symbol(code, "KRX")
                average_cost = (principal / qty).quantize(Decimal("0.0001"))
                data.append(
                    ParsedHolding(
                        asset_type=AssetType.pension,
                        symbol=symbol,
                        market_symbol=symbol,
                        name=item_name or symbol,
                        category=group,
                        owner_name=owner,
                        account_name=company,
                        quantity=qty,
                        average_cost=average_cost,
                        currency="KRW",
                        source_ref=f"excel:{name}:{row_idx}",
                    )
                )
            else:
                symbol = self._build_cash_symbol(item_name or "pension", code, row_idx, "PENSION")
                data.append(
                    ParsedHolding(
                        asset_type=AssetType.pension,
                        symbol=symbol,
                        market_symbol=symbol,
                        name=item_name or symbol,
                        category=group,
                        owner_name=owner,
                        account_name=company,
                        quantity=Decimal("1"),
                        average_cost=principal.quantize(Decimal("0.0001")),
                        currency="KRW",
                        source_ref=f"excel:{name}:{row_idx}",
                    )
                )
        return data

    def _parse_real_estate_holdings(self, wb, issues: list[ImportIssue]) -> list[ParsedHolding]:
        data: list[ParsedHolding] = []
        name = "1) 부동산"
        if name not in wb.sheetnames:
            return data
        ws = wb[name]
        seen: set[tuple[str, str, Decimal]] = set()
        upper = self._sheet_row_upper_bound_with_issue(ws, issues=issues, sheet_name=name)
        for row_idx, row_values in self._iter_rows_values(ws, min_row=12, max_row=upper, max_col=10):
            district = self._text(self._row_value(row_values, 2))
            neighborhood = self._text(self._row_value(row_values, 3))
            complex_name = self._text(self._row_value(row_values, 4))
            sell_price = self._to_decimal(self._row_value(row_values, 10))
            if not complex_name or sell_price is None or sell_price <= 0:
                continue
            if district and "합계" in district:
                continue
            normalized = sell_price.quantize(Decimal("0.0001"))
            if normalized < Decimal("1000000"):
                normalized = (normalized * Decimal("10000")).quantize(Decimal("0.0001"))
            account_name = " / ".join([part for part in [district, neighborhood] if part]) or "부동산"
            key = (complex_name, account_name, normalized)
            if key in seen:
                continue
            seen.add(key)
            symbol = self._build_cash_symbol(f"{complex_name}-{row_idx}", account_name, row_idx, "RE")
            data.append(
                ParsedHolding(
                    asset_type=AssetType.real_estate,
                    symbol=symbol,
                    market_symbol=symbol,
                    name=complex_name,
                    category="부동산",
                    owner_name=None,
                    account_name=account_name,
                    quantity=Decimal("1"),
                    average_cost=normalized,
                    currency="KRW",
                    source_ref=f"excel:{name}:{row_idx}",
                )
            )
        return data

    def _upsert_categories(
        self,
        db: Session,
        household_id: str,
        rows: list[tuple[FlowType, str, str]],
    ) -> tuple[dict[tuple[FlowType, str, str], Category], int]:
        existing = db.scalars(select(Category).where(Category.household_id == household_id)).all()
        by_key = {(item.flow_type, item.major, item.minor): item for item in existing}
        inserted = 0
        for idx, (flow_type, major, minor) in enumerate(rows):
            key = (flow_type, major, minor)
            if key in by_key:
                continue
            entity = Category(
                household_id=household_id,
                flow_type=flow_type,
                major=major,
                minor=minor,
                sort_order=idx + 1,
            )
            db.add(entity)
            db.flush()
            by_key[key] = entity
            inserted += 1
        return by_key, inserted

    def _apply_transactions(
        self,
        db: Session,
        household_id: str,
        category_map: dict[tuple[FlowType, str, str], Category],
        rows: list[ParsedTransaction],
    ) -> tuple[int, int]:
        existing_transactions = db.scalars(
            select(Transaction).where(Transaction.household_id == household_id)
        ).all()
        existing_sources = {
            item
            for item in (entity.source_ref for entity in existing_transactions)
            if item
        }
        existing_hash_counts: dict[str, int] = {}
        category_ids = {
            str(entity.category_id or "").strip()
            for entity in existing_transactions
            if str(entity.category_id or "").strip()
        }
        categories_by_id: dict[str, Category] = {}
        if category_ids:
            categories_by_id = {
                str(item.id): item
                for item in db.scalars(select(Category).where(Category.id.in_(category_ids))).all()
            }
        for entity in existing_transactions:
            category = categories_by_id.get(str(entity.category_id or "").strip())
            if category is None:
                continue
            dedupe_hash = self._transaction_dedupe_hash(
                occurred_on=entity.occurred_on,
                flow_type=entity.flow_type,
                major=category.major,
                minor=category.minor,
                amount=entity.amount,
                memo=entity.memo or "",
                owner_name=entity.owner_name,
            )
            existing_hash_counts[dedupe_hash] = int(existing_hash_counts.get(dedupe_hash, 0)) + 1
        added = 0
        skipped = 0
        for row in rows:
            existing_same_hash = int(existing_hash_counts.get(row.dedupe_hash, 0))
            if row.source_ref in existing_sources or row.dedupe_ordinal <= existing_same_hash:
                skipped += 1
                continue
            key = (row.flow_type, row.major, row.minor)
            category = category_map.get(key)
            if category is None:
                category = Category(
                    household_id=household_id,
                    flow_type=row.flow_type,
                    major=row.major,
                    minor=row.minor,
                    sort_order=len(category_map) + 1,
                )
                db.add(category)
                db.flush()
                category_map[key] = category

            db.add(
                Transaction(
                    household_id=household_id,
                    category_id=category.id,
                    flow_type=row.flow_type,
                    occurred_on=row.occurred_on,
                    amount=row.amount,
                    currency="KRW",
                    memo=row.memo,
                    owner_user_id=(
                        linked_user.id
                        if (linked_user := find_unique_household_member_by_display_name(
                            db,
                            household_id=household_id,
                            display_name=row.owner_name,
                        ))
                        is not None
                        else None
                    ),
                    owner_name=row.owner_name,
                    source_ref=row.source_ref,
                )
            )
            existing_sources.add(row.source_ref)
            existing_hash_counts[row.dedupe_hash] = max(existing_same_hash, row.dedupe_ordinal)
            added += 1
        return added, skipped

    def _apply_holdings(self, db: Session, household_id: str, rows: list[ParsedHolding]) -> tuple[int, int, list[ImportIssue]]:
        existing = db.scalars(select(Holding).where(Holding.household_id == household_id)).all()
        invalid_owner_names: set[str] = set()
        ambiguous_owner_names: set[str] = set()
        apply_issues: list[ImportIssue] = []

        def resolve_owner(value: str | None) -> tuple[str | None, str]:
            normalized_owner = self._normalize_holder_text(value)
            if not normalized_owner:
                return None, ""
            linked_user = find_unique_household_member_by_display_name(
                db,
                household_id=household_id,
                display_name=normalized_owner,
            )
            if linked_user is not None:
                return str(linked_user.id), normalized_owner
            duplicate_member_count = int(
                db.scalar(
                    select(func.count())
                    .select_from(HouseholdMember)
                    .join(User, User.id == HouseholdMember.user_id)
                    .where(
                        HouseholdMember.household_id == household_id,
                        func.lower(User.display_name) == normalized_owner.lower(),
                    )
                )
                or 0
            )
            if duplicate_member_count > 1:
                ambiguous_owner_names.add(normalized_owner)
            else:
                invalid_owner_names.add(normalized_owner)
            return None, normalized_owner

        for item in existing:
            item.version = self._normalize_version(item.version)
        by_key = {
            self._holding_key(
                item.asset_type,
                item.market_symbol,
                item.owner_user_id,
                item.owner_name,
                item.account_name,
            ): item
            for item in existing
        }
        by_source_ref = {
            item.source_ref: item
            for item in existing
            if item.source_ref
        }
        added = 0
        updated = 0
        for row in rows:
            owner_user_id, normalized_owner = resolve_owner(row.owner_name)
            normalized_account = self._normalize_holder_text(row.account_name)
            key = self._holding_key(
                row.asset_type,
                row.market_symbol,
                owner_user_id,
                normalized_owner,
                normalized_account,
            )
            entity = by_source_ref.get(row.source_ref) if row.source_ref else None
            if entity is None:
                entity = by_key.get(key)
            if entity is None:
                entity = Holding(
                    household_id=household_id,
                    asset_type=row.asset_type,
                    symbol=row.symbol,
                    market_symbol=row.market_symbol,
                    name=row.name,
                    category=row.category,
                    owner_user_id=owner_user_id,
                    owner_name=normalized_owner,
                    account_name=normalized_account,
                    quantity=row.quantity,
                    average_cost=row.average_cost,
                    currency=row.currency,
                    source_ref=row.source_ref,
                )
                db.add(entity)
                by_key[key] = entity
                if row.source_ref:
                    by_source_ref[row.source_ref] = entity
                added += 1
                continue

            changed = False
            old_key = self._holding_key(
                entity.asset_type,
                entity.market_symbol,
                entity.owner_user_id,
                entity.owner_name,
                entity.account_name,
            )
            for attr in (
                "symbol",
                "market_symbol",
                "name",
                "category",
                "owner_name",
                "account_name",
                "quantity",
                "average_cost",
                "currency",
                "source_ref",
            ):
                value = getattr(row, attr)
                if attr in {"owner_name", "account_name"}:
                    if attr == "owner_name":
                        linked_owner_user_id, value = resolve_owner(value)
                        if entity.owner_user_id != linked_owner_user_id:
                            entity.owner_user_id = linked_owner_user_id
                            changed = True
                    else:
                        value = self._normalize_holder_text(value)
                if getattr(entity, attr) != value:
                    setattr(entity, attr, value)
                    changed = True
            if changed:
                entity.version = self._normalize_version(entity.version) + 1
                updated += 1
            new_key = self._holding_key(
                entity.asset_type,
                entity.market_symbol,
                entity.owner_user_id,
                entity.owner_name,
                entity.account_name,
            )
            if old_key != new_key and by_key.get(old_key) is entity:
                by_key.pop(old_key, None)
            by_key[new_key] = entity
            if entity.source_ref:
                by_source_ref[entity.source_ref] = entity
        for owner_name in sorted(invalid_owner_names):
            apply_issues.append(
                ImportIssue(
                    severity="warning",
                    code="HOLDING_OWNER_NOT_MEMBER",
                    message="가계 구성원이 아닌 보유자는 legacy owner_name 으로 보존했습니다.",
                    detail={"owner_name": owner_name},
                )
            )
        for owner_name in sorted(ambiguous_owner_names):
            apply_issues.append(
                ImportIssue(
                    severity="warning",
                    code="HOLDING_OWNER_AMBIGUOUS",
                    message="동일한 표시 이름 구성원이 여러 명이라 owner_user_id 없이 legacy owner_name 만 보존했습니다.",
                    detail={"owner_name": owner_name},
                )
            )
        return added, updated, apply_issues

    @staticmethod
    def _holding_key(
        asset_type: AssetType,
        market_symbol: str,
        owner_user_id: str | None,
        owner_name: str | None,
        account_name: str | None,
    ) -> tuple[AssetType, str, str, str]:
        return (
            asset_type,
            str(market_symbol or "").strip().upper(),
            str(owner_user_id or "").strip() or str(owner_name or "").strip(),
            str(account_name or "").strip(),
        )

    @staticmethod
    def _normalize_holder_text(value: str | None) -> str:
        return str(value or "").strip()

    @staticmethod
    def _decimal_key(value: Decimal) -> str:
        normalized = format(value.normalize(), "f")
        trimmed = normalized.rstrip("0").rstrip(".")
        if trimmed in {"", "-0"}:
            return "0"
        return trimmed

    @classmethod
    def _transaction_dedupe_hash(
        cls,
        *,
        occurred_on: date,
        flow_type: FlowType,
        major: str,
        minor: str,
        amount: Decimal,
        memo: str,
        owner_name: str | None,
    ) -> str:
        payload = "|".join(
            [
                occurred_on.isoformat(),
                str(flow_type.value if isinstance(flow_type, FlowType) else flow_type).strip().lower(),
                str(major or "").strip(),
                str(minor or "").strip(),
                cls._decimal_key(amount),
                str(memo or "").strip(),
                str(owner_name or "").strip(),
            ]
        )
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]

    @staticmethod
    def _guess_flow_type(text: str) -> FlowType:
        src = text.replace(" ", "").lower()
        if "수입" in src or "급여" in src or "보너스" in src:
            return FlowType.income
        if "이체" in src:
            return FlowType.transfer
        if "저축" in src or "투자" in src or "주식" in src or "연금" in src:
            return FlowType.investment
        return FlowType.expense

    @staticmethod
    def _normalize_stock_symbol(code: str, market: str | None) -> str:
        text = code.strip().upper()
        normalized_market = (market or "").upper()
        is_krx_like_code = bool(re.fullmatch(r"[0-9A-Z]{6}", text))
        if normalized_market in {"KRX", "KOSDAQ"} and "." not in text and is_krx_like_code:
            return f"{text}.KR"
        return text

    @staticmethod
    def _build_cash_symbol(name: str, account: str | None, row_idx: int, prefix: str) -> str:
        digits = "".join(ch for ch in (account or "") if ch.isdigit())
        token = digits[-12:] if len(digits) >= 6 else re.sub(r"[^A-Za-z0-9가-힣]+", "-", name).strip("-").upper()
        if not token:
            token = f"ROW{row_idx}"
        return f"{prefix}-{token}"[:40]

    @staticmethod
    def _extract_owner_name(major: str | None, minor: str | None, memo: str | None) -> str | None:
        for text in [memo, minor, major]:
            if not text:
                continue
            value = text.strip()
            if value.startswith("댕"):
                return "댕"
            if value.startswith("찌"):
                return "찌"
            if "댕:" in value or "댕 -" in value:
                return "댕"
            if "찌:" in value or "찌 -" in value:
                return "찌"
        return None

    @staticmethod
    def _to_date(value: object) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            try:
                return datetime.fromisoformat(text).date()
            except ValueError:
                return None
        return None

    @staticmethod
    def _to_decimal(value: object) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        if isinstance(value, str):
            text = value.strip().replace(",", "")
            if text in {"", "-", "N/A", " "}:
                return None
            try:
                return Decimal(text)
            except InvalidOperation:
                return None
        return None

    @staticmethod
    def _text(value: object) -> str | None:
        if value is None:
            return None
        text = str(value).strip().replace("\n", " ")
        return text or None

    @staticmethod
    def _row_value(row_values: tuple[object, ...], column_index: int) -> object | None:
        offset = column_index - 1
        if offset < 0 or offset >= len(row_values):
            return None
        return row_values[offset]

    @staticmethod
    def _iter_rows_values(
        ws,
        *,
        min_row: int,
        max_row: int,
        max_col: int,
    ):
        for row_idx, row_values in enumerate(
            ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col, values_only=True),
            start=min_row,
        ):
            yield row_idx, row_values

    @staticmethod
    def _normalize_version(value: object) -> int:
        try:
            parsed = int(value)
            return parsed if parsed > 0 else 1
        except (TypeError, ValueError):
            return 1

    def _sheet_row_upper_bound_with_issue(self, ws, *, issues: list[ImportIssue], sheet_name: str) -> int:
        row_cap = max(1, int(settings.import_max_rows_per_sheet))
        upper = self._sheet_row_upper_bound(ws, row_cap)
        max_row = getattr(ws, "max_row", None)
        if isinstance(max_row, int) and max_row > row_cap:
            issues.append(
                ImportIssue(
                    severity="warning",
                    code="IMPORT_ROW_LIMIT_TRUNCATED",
                    message=f"행 제한({row_cap})을 초과한 데이터는 가져오지 않습니다.",
                    sheet=sheet_name,
                    row=row_cap + 1,
                )
            )
        return upper

    @staticmethod
    def _sheet_row_upper_bound(ws, cap: int) -> int:
        max_row = getattr(ws, "max_row", None)
        if isinstance(max_row, int) and max_row > 0:
            return min(max_row, cap)
        return cap

    @staticmethod
    def _sheet_col_upper_bound(ws, cap: int) -> int:
        max_col = getattr(ws, "max_column", None)
        if isinstance(max_col, int) and max_col > 0:
            return min(max_col, cap)
        return cap
